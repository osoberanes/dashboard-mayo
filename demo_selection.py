import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random

def create_demo_with_selections():
    """Create a demo version with some contacts pre-selected"""
    
    print("Creating demo version with sample selections...")
    
    # Read the created file
    df = pd.read_excel("C:/Users/consuladscrito/Documents/invitados.xlsx", sheet_name="Directorio Contactos")
    
    print(f"Total contacts available: {len(df)}")
    
    # Randomly select about 10-15 contacts for demo
    n_selections = min(15, len(df))
    selected_indices = random.sample(range(len(df)), n_selections)
    
    # Mark selected contacts
    df.loc[selected_indices, 'FN25'] = True
    
    # Create filtered dataframe for FN25 sheet
    fn25_df = df[df['FN25'] == True][['nombre', 'apellido', 'cargo', 'correo_electronico']].copy()
    
    print(f"Selected {len(fn25_df)} contacts for demo")
    
    # Load the existing workbook to preserve formatting
    wb = load_workbook("C:/Users/consuladscrito/Documents/invitados.xlsx")
    
    # Update the main sheet with new FN25 values
    ws1 = wb["Directorio Contactos"]
    
    # Update FN25 column (column A) with checkmarks
    for idx, row in df.iterrows():
        cell_row = idx + 2  # +2 because of header and 0-based indexing
        fn25_value = "☑" if row['FN25'] else "☐"
        ws1[f"A{cell_row}"] = fn25_value
    
    # Update the FN25 sheet with selected contacts
    ws2 = wb["Invitados FN25"]
    
    # Clear existing data
    ws2.delete_rows(1, ws2.max_row)
    
    # Add new data
    for r_idx, row in enumerate(dataframe_to_rows(fn25_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    
    # Save with demo suffix
    demo_filename = "C:/Users/consuladscrito/Documents/invitados_demo.xlsx"
    wb.save(demo_filename)
    wb.close()
    
    print(f"Demo file created: {demo_filename}")
    print("\\nDemo includes:")
    for idx, row in fn25_df.iterrows():
        print(f"- {row['nombre']} {row['apellido']} ({row['cargo']})")
    
    return demo_filename

if __name__ == "__main__":
    create_demo_with_selections()