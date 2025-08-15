import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

def clean_and_consolidate_contacts():
    """Process and consolidate contact data from Excel file"""
    
    # Read all sheets
    file_path = 'c:/Users/consuladscrito/Documents/Contactos_SPM_AGO_7_2025F4.xlsx'
    excel_file = pd.ExcelFile(file_path)
    
    print("Processing all contact sheets...")
    all_contacts = []
    
    for sheet_name in excel_file.sheet_names:
        print(f"Processing sheet: {sheet_name}")
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df['FUENTE'] = sheet_name
            all_contacts.append(df)
        except Exception as e:
            print(f"Error reading {sheet_name}: {e}")
    
    # Combine all dataframes
    print("Combining all contacts...")
    combined_df = pd.concat(all_contacts, ignore_index=True, sort=False)
    print(f"Total records before processing: {len(combined_df)}")
    
    # Standardize column names - map to consistent names
    column_mapping = {
        'NOMBRE': 'nombre',
        'APELLIDO': 'apellido', 
        'CARGO': 'cargo',
        'CORREO ELECTRONICO': 'correo_electronico',
        'OTRO CORREO ELECTRONICO': 'correo_electronico_2',
        'ORGANISMO': 'organismo',
        'TELEFONO': 'telefono',
        'TELEFONO OFICINA': 'telefono_oficina',
        'DIRECCION': 'direccion',
        'CIUDAD': 'ciudad',
        'ESTADO': 'estado',
        'PAIS': 'pais',
        'C. POSTAL': 'codigo_postal',
        'WEBSITE': 'website',
        'TIT': 'titulo',
        'IDIOMA': 'idioma',
        'FUENTE': 'fuente'
    }
    
    # Select and rename relevant columns
    relevant_columns = []
    for old_col, new_col in column_mapping.items():
        if old_col in combined_df.columns:
            relevant_columns.append(old_col)
    
    # Create cleaned dataframe with only relevant columns
    cleaned_df = combined_df[relevant_columns].copy()
    cleaned_df = cleaned_df.rename(columns=column_mapping)
    
    # Clean data
    print("Cleaning data...")
    
    # Remove rows where both name and lastname are empty
    cleaned_df = cleaned_df.dropna(subset=['nombre', 'apellido'], how='all')
    
    # Clean text fields
    text_columns = ['nombre', 'apellido', 'cargo', 'organismo', 'direccion', 'ciudad']
    for col in text_columns:
        if col in cleaned_df.columns:
            cleaned_df[col] = cleaned_df[col].astype(str).str.strip()
            cleaned_df[col] = cleaned_df[col].replace(['nan', 'None', ''], np.nan)
    
    # Clean email fields
    email_columns = ['correo_electronico', 'correo_electronico_2']
    for col in email_columns:
        if col in cleaned_df.columns:
            cleaned_df[col] = cleaned_df[col].astype(str).str.strip().str.lower()
            cleaned_df[col] = cleaned_df[col].replace(['nan', 'none', ''], np.nan)
            # Basic email validation
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            invalid_emails = ~cleaned_df[col].str.match(email_pattern, na=True)
            cleaned_df.loc[invalid_emails, col] = np.nan
    
    # Create full name for duplicate detection
    cleaned_df['nombre_completo'] = (
        cleaned_df['nombre'].fillna('') + ' ' + cleaned_df['apellido'].fillna('')
    ).str.strip()
    
    # Remove duplicates based on full name and primary email
    print("Removing duplicates...")
    
    # Remove exact duplicates
    initial_count = len(cleaned_df)
    cleaned_df = cleaned_df.drop_duplicates(subset=['nombre_completo', 'correo_electronico'], keep='first')
    duplicates_removed = initial_count - len(cleaned_df)
    print(f"Removed {duplicates_removed} duplicate records")
    
    # Add FN25 column (checkbox for future selection)
    cleaned_df['FN25'] = False
    
    # Reorder columns for better presentation
    column_order = [
        'FN25', 'titulo', 'nombre', 'apellido', 'cargo', 'organismo',
        'correo_electronico', 'correo_electronico_2', 'telefono', 'telefono_oficina',
        'direccion', 'ciudad', 'estado', 'pais', 'codigo_postal',
        'website', 'idioma', 'fuente'
    ]
    
    # Only include columns that exist
    final_columns = [col for col in column_order if col in cleaned_df.columns]
    cleaned_df = cleaned_df[final_columns]
    
    # Remove the helper column
    if 'nombre_completo' in cleaned_df.columns:
        cleaned_df = cleaned_df.drop('nombre_completo', axis=1)
    
    print(f"Final records count: {len(cleaned_df)}")
    return cleaned_df

def create_formatted_excel(df, filename):
    """Create a beautifully formatted Excel file"""
    
    print("Creating formatted Excel file...")
    
    # Create workbook with two sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Directorio Contactos"
    ws2 = wb.create_sheet("Invitados FN25")
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"), 
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Sheet 1: All contacts
    print("Formatting main contacts sheet...")
    
    # Add data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.font = data_font
            cell.alignment = data_alignment
            
            # Format header row
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Special formatting for FN25 column (checkbox)
            if c_idx == 1 and r_idx > 1:  # FN25 is first column
                cell.value = "☐" if not value else "☑"
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for column_cells in ws1.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(length + 2, 50)  # Cap at 50 characters
        ws1.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    # Create table for better formatting
    from openpyxl.utils import get_column_letter
    max_col_letter = get_column_letter(ws1.max_column)
    table1 = Table(
        displayName="ContactosTable",
        ref=f"A1:{max_col_letter}{ws1.max_row}"
    )
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws1.add_table(table1)
    
    # Sheet 2: FN25 filtered contacts (dynamic table)
    print("Creating FN25 filtered sheet...")
    
    # Filter data where FN25 is True
    fn25_df = df[df['FN25'] == True][['nombre', 'apellido', 'cargo', 'correo_electronico']].copy()
    
    if len(fn25_df) == 0:
        # Create sample header if no data
        fn25_df = pd.DataFrame({
            'nombre': ['(Selecciona contactos en la hoja principal)'],
            'apellido': [''],
            'cargo': [''],
            'correo_electronico': ['']
        })
    
    # Add data to second worksheet
    for r_idx, row in enumerate(dataframe_to_rows(fn25_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.font = data_font
            cell.alignment = data_alignment
            
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    
    # Auto-adjust column widths for sheet 2
    for column_cells in ws2.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(length + 2, 50)
        ws2.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    # Create table for sheet 2
    if ws2.max_row > 1:
        max_col_letter2 = get_column_letter(ws2.max_column)
        table2 = Table(
            displayName="InvitadosTable",
            ref=f"A1:{max_col_letter2}{ws2.max_row}"
        )
        table2.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium15",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws2.add_table(table2)
    
    # Save workbook
    wb.save(filename)
    print(f"Excel file saved as: {filename}")

def main():
    """Main execution function"""
    try:
        # Process contacts
        cleaned_contacts = clean_and_consolidate_contacts()
        
        # Create output filename
        output_filename = "C:/Users/consuladscrito/Documents/invitados.xlsx"
        
        # Create formatted Excel
        create_formatted_excel(cleaned_contacts, output_filename)
        
        print("\\nProcess completed successfully!")
        print(f"File created: {output_filename}")
        print(f"Total contacts: {len(cleaned_contacts)}")
        print("\\nNext steps:")
        print("1. Open the Excel file")
        print("2. Mark contacts as selected by changing FN25 to TRUE")
        print("3. The 'Invitados FN25' sheet will automatically show selected contacts")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()